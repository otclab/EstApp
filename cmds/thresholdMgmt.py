#!/usr/bin/python
#-*- coding: utf-8 -*-

import sys
from dataclasses import dataclass
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side
from common import openCard


def ThresholdCmd(args, port, throughput_limit) :
  """
    EstApp : Umbrales de Tensión de los Taps
    ===========================================

    Tiene cinco formas de invocarlo :
      (a) Presenta una tabla con la lista de los taps activos :

          >> EstApp.py -u
          o
          >> EstApp.py -u  ?

      (b) Modifica el número de cambios, al valor definido por
          Nuevo_Valor_del_Numero_de_Cambios :

          >> EstApp.py -u total Nuevo_Valor_del_Numero_de_Cambios

      (c) Modifica el valor de un umbral, cuyo número de orden de menor a mayor
          es Nro_de_Orden y cuya posición superior o inferior es indicada por
          Pos con el valor dado por Nuevo_Valor_del_Umbral :

          >> EstApp.py -u  Nro_de_Orden  Pos  Nuevo_Valor_del_Umbral

          El número de orden se indica en forma abreviada, es decir puede ser
          '1ro', '2do', '3ro', '4to', '5to', '6to', '7mo', '8vo', '9no', '10mo',
          '11avo' o '12avo'

          La posición se indica por las palabras 'sup' o 'superior' para el
          valor del umbral de cambio al siguiente tap e 'inf' o'inferior' para
          el valor del umbral al tap anterior.

       (d) Modifica los umbrales definidos en un archivo de una hoja de calculo
          Excel.

          >> EstApp.py -u -r Nombre_del_archivo

          El nombre del archivo debe incluir la ruta, si es que no se encuentra
          en el directorio actual. Si contiene espacios deberá incluirse entre
          comillas (dobles). En la hoja de Excel los umbrales de cada tap se
          definen a partir de la segunda fila, en la segunda columna (B) el
          umbral superior y en la tercera (C) el inferior.


       (e) Lee los umbrales definidos en la tarjeta y los guarda en una hoja de
           calculo Excel.

          >> EstApp.py -u -w Nombre_del_archivo

          El nombre del archivo debe incluir la ruta, si es que no se encuentra
          en el directorio actual. Si contiene espacios deberá incluirse entre
          comillas (dobles). En la hoja de Excel los umbrales de cada tap se
          definen a partir de la segunda fila, en la segunda columna (B) el
          umbral superior y en la tercera (C) el inferior.


  """

  num_tap = ['1ro', '2do', '3ro', '4to', '5to', '6to',
                               '7mo', '8vo', '9no', '10mo', '11avo', '12avo']

  card = openCard(port, throughput_limit)

  if (len(args) < 3) or (args[2] == u'?') :
    print(card.threshold)

  elif args[2] == 'total' :
    if (len(args) < 4) or (args[3] == u'?') :
       print('El Número Total de Taps es : %d' %len(card.threshold))
    else :
      try :
        val = int(u' '.join(args[3:]))
      except :
        print('Error : "%s" no es un número.' % args[3:])
        card.close()
        sys.exit(1)

      try :
        card.threshold.len = val
      except ValueError as e :
        print(e.message)
        card.close()
        sys.exit(1)

      print(u'El Número Total de Taps (operativos) se '                     \
                                        u'reajusto a %d' %len(card.threshold))
  elif args[2] in num_tap :
    if not (num_tap.index(args[2]) < len(card.threshold)) :
      print(u'Error : El tap excede el número de taps operativos.')
      card.close()
      sys.exit(1)

    if len(args) < 4 :
      print ('Los Umbrales del tap %s  son :\n  ' % args[2],)
      print (card.threshold[num_tap.index(args[2])])

    else :
      pos = [u'sup', u'superior', u'inf', u'inferior']
      if not args[3].lower() in pos :
        print('Error : %s no indica la posición del umbral '               \
                                      '(sup(erior) o inf(erior)).' % args[3])
        card.close()
        sys.exit(1)

      tap_idx = num_tap.index(args[2])
      if len(args) < 5 :
        print('El Umbral %s del %s tap es %s' %(
                   pos[pos.index(args[3][:3])+1].capitalize(), args[2],
                               getattr(card.threshold[tap_idx], args[3][:3])))
      else :
        try :
          val = float(args[4])
          setattr(card.threshold[tap_idx], args[3][:3], val)
        except Exception as e:
          print("Error : %s no es un número o esta fuera de rango." % args[4])
          card.close()
          sys.exit(1)

        print('Se reajusto el Umbral %s del %s tap a %s' %(
                   pos[pos.index(args[3][:3])+1].capitalize(), args[2],
                               getattr(card.threshold[tap_idx], args[3][:3])))

  elif args[2] == '-w' :
    print(f'args : {args}')
    if len(args) > 3 :
      write_xls(card, args[3])

    else :
      print(f'Error : Falta el nombre de la hoja Excel.')
      card.close()
      sys.exit(1)

  elif args[2] == '-r' :
    if len(args) > 3 :
      read_xls(card, args[3])

    else :
      print(f'Error : Falta el nombre de la hoja Excel.')
      card.close()
      sys.exit(1)


  else :
    print(f'Error : {args[2]} no es el sub-comando "total", "r" o "w" ni '    \
                                                      f'el índice de un tap.')
    print('Pruebe la opción -h -u para ver los detalles.')
    card.close()
    sys.exit(1)

  card.close()


@dataclass
class Threshold:
    sup : float
    inf : float

    def __post_init__(self) :
        if self.sup is None or self.inf is None :
            raise ValueError(f'Se trato de asignar un umbral incompleto.')

class ThresholdList(list) :
    def __init__(self, filename = None) :
        wb = load_workbook(filename = filename)
        ws = wb.active

        # Se validan los umbrales
        super().__init__()
        for n, row in enumerate(ws[2:7]) :
            sup, inf  = row[1].value, row[2].value

            if sup is None and inf is None :
                break

            if sup is None :
                raise ValueError(f'Falta definir el umbral superior del tap {n+1}.')

            if not isinstance(sup, (int, float)) :
                raise ValueError(f'El umbral superior del tap {n+1} no es un número.')

            if inf is None :
                raise ValueError(f'Falta definir el umbral inferior del tap {n+1}.')

            if not isinstance(inf, (int, float)) :
                raise ValueError(f'El umbral inferior del tap { n+1} no es un número.')

            if sup <= inf :
                raise ValueError('Los umbrales del tap {n+1} estan invertidos.')

            if super().__len__() > 1 :
                if self[-1].sup < inf :
                    msg = f'Los taps {n} y {n+1} no se traslapan.\n' +                                \
                          f'(El umbral superior del tap {n} es menor que el inferior del tap {n+1})'
                    raise ValueError(msg)

            super().append(Threshold(sup, inf))


def read_xls(card, xls_name) :
  try :
    threshold_list = ThresholdList(filename = xls_name)
  except ValueError as e :
    print(f'Error : {e.args[0]}')
    card.close()
    sys.exit(1)

  for n, th in enumerate(threshold_list) :
     setattr(card.threshold[n], 'sup', th.sup)
     setattr(card.threshold[n], 'inf', th.inf)

  card.threshold.len = len(threshold_list)

  print(f'Los umbrales se guardaron en \'{xls_name}\'')


def write_xls(card, xls_name) :
  accounting_format = '_ * #,##0.00_ ;_ * \\-#,##0.00_ ;_ * "-"??_ ;_ @_ '
  thin  = Side(border_style='thin', color='000000')
  thick = Side(border_style='thick', color='000000')

  wb = Workbook()
  ws = wb.active

  ws.title = 'Asignación de Taps'
  ws['A1'], ws['B1'], ws['C1'] = '# Tap', 'Umbral\nAlto', 'Umbral\nBajo'

  for c in 'A1 B1 C1'.split() :
    ws[c].alignment = Alignment(wrapText='True',
                                      horizontal='center', vertical='center')

  for c in 'A1 B1 C1'.split() :
    ws[c].border = Border(top=thick, left=thick, right=thick, bottom=thick)


  for n, th in enumerate(card.threshold[:card.threshold.len]) :
    ws[f'A{n+2}'], ws[f'B{n+2}'], ws[f'C{n+2}'] = n+1, th.sup, th.inf
    ws[f'A{n+2}'].alignment = Alignment(wrapText='True',
                                      horizontal='center', vertical='center')
    ws[f'B{n+2}'].number_format = accounting_format
    ws[f'C{n+2}'].number_format = accounting_format

    bottom = thick if n == (card.threshold.len - 1) else thin
    ws[f'A{n+2}'].border = Border(top=thin , left=thick, right=thick, bottom=bottom)
    ws[f'B{n+2}'].border = Border(top=thin , left=thick, right=thin , bottom=bottom)
    ws[f'C{n+2}'].border = Border(top=thin , left=thin , right=thick, bottom=bottom)



  wb.save(xls_name)


